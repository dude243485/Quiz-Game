import tkinter as tk
import time
from tkinter import messagebox, Toplevel
from openpyxl import load_workbook

excel_sheet = 'quiz.xlsx'
questions = []
current_question = 0
score = 0
coins = 0
time_remaining = 60  # Total quiz time in seconds
timer_running = False
question_label = None
selected_button = None
buttons = []
is_paused = False
pause_window = None
selected_options = {}
previous_answers = {}
answered_questions = set()

def question_section():
    global questions
    try:
        workbook = load_workbook(excel_sheet)
        sheet = workbook.active
        questions.clear()
        # getting questions from the data sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            question = {
                "text": row[0],
                "options": [row[1], row[2], row[3], row[4]],
                "answer": row[6]  # Correct answer is now from column G
            }
            questions.append(question)
    except FileNotFoundError as e:
        messagebox.showerror("Error", f"File not found: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load questions: {e}")
# timer function to countdown the time
def start_quiz_timer():
    global time_remaining, timer_running
    if not timer_running:
        timer_running = True
        countdown()

def countdown():
    global time_remaining, timer_running
    if time_remaining > 0:
        timer_label.config(text=f"Time Left: {time_remaining}s")
        time_remaining -= 1
        app.after(1000, countdown)
    else:
        timer_running = False
        messagebox.showinfo("Time's Up!", "The quiz has ended.")
        app.quit()
# this is the function responsible for viewing the questions
def showquestion():
    global current_question, questions, selected_button, timer_label
    if is_paused:
        return
    if current_question < len(questions):
        question = questions[current_question]
        question_label.config(text=question["text"])
        
        for btn in buttons:
            btn.config(borderwidth=0, relief="flat", bg="#1581B4", text="")
        
        selected_button = None
        # option buttons
        for i, option in enumerate(question["options"]):
            buttons[i].config(text=option, command=lambda opt=option, btn=buttons[i]: handle_selection(opt, btn))
            if selected_options.get(current_question) == option:
                select_option(buttons[i])
                # timer config
        timer_label.config(text=f"Time Left: {time_remaining}s")
    else:
        messagebox.showinfo("Quiz Completed", f"Your score: {score}/{len(questions)}")
        app.quit()
# For clicking each button to select an option
def handle_selection(option, btn):
    global selected_options, previous_answers, current_question, score, coins, answered_questions
    if current_question >= len(questions):
        return
    
    correct_answer = questions[current_question]["answer"]
    previous_answer = previous_answers.get(current_question)
    if option == correct_answer:
        score += 1
        coins += 5
        # if previous_answer != correct_answer:
        #     score += 1
        #     coins += 5
    elif previous_answer == correct_answer:
        score -= 1
        coins -= 5
    
    selected_options[current_question] = option
    previous_answers[current_question] = option
    select_option(btn)
    coin_score_increase()

# next question
def next_question():
    global current_question
    if current_question < len(questions) - 1:
        current_question += 1
        showquestion()
# the coins and score update
def coin_score_increase():
    score_label.config(text=f"Score: {score}")
    coin_label.config(text=f"Coins: {coins}")
    app.update_idletasks()
# previous question
def previous_question():
    global current_question
    if current_question > 0:
        current_question -= 1
        showquestion()
    else:
        messagebox.showwarning("Navigation", "This is the first question!")
# select option button
def select_option(btn):
    global selected_button
    if selected_button:
        selected_button.config(borderwidth=0, relief="flat", bg="#1581B4")
    btn.config(borderwidth=2, relief="solid", bg="#EDF7F7")
    selected_button = btn
# pause function
def pause():
    global is_paused, pause_window
    is_paused = True
    pause_window = Toplevel(app)
    pause_window.geometry("200x300")
    pause_window.resizable(False, False)
    tk.Label(pause_window, text="Paused", font=("Montserrat", 18)).pack(pady=10)
    tk.Button(pause_window, text="Resume", command=resume_quiz, font=("Montserrat", 16), bg="#1581B4", fg="#EDF7F7").pack(pady=10)
    tk.Button(pause_window, text="Restart", command=restart_quiz, font=("Montserrat", 16), bg="pink", fg="#EDF7F7").pack(pady=10)
    tk.Button(pause_window, text="Quit", command=quit_quiz, font=("Montserrat", 16), bg="pink", fg="#EDF7F7").pack(pady=10)
    pause_window.transient(app)
    pause_window.grab_set()
# resume quiz
def resume_quiz():
    global is_paused, pause_window
    is_paused = False
    pause_window.destroy()
    showquestion()
# restart quiz
def restart_quiz():
    global current_question, score, coins, selected_button, pause_window, selected_options, previous_answers, answered_questions, time_remaining, timer_running
    user_response = messagebox.askyesno('Restart', 'Are you sure you want to restart?')
    if user_response:
        current_question, score, coins = 0, 0, 0
        time_remaining = 60
        timer_running = False
        selected_button = None
        selected_options.clear()
        previous_answers.clear()
        answered_questions.clear()
        if pause_window:
            pause_window.destroy()
        question_section()
        coin_score_increase()
        showquestion()
        start_quiz_timer()

def quit_quiz():
    if messagebox.askyesno("Quit", "Are you sure you want to exit?"):
        app.quit()

# UI components
app = tk.Tk()
timer_label = tk.Label(app, text=f"Time Left: {time_remaining}s", font=("Montserrat", 14), fg="red")
timer_label.pack(pady=5)
start_quiz_timer()

score_label = tk.Label(app, text="Score: 0", font=("Montserrat", 16), fg="#1581B4")
score_label.pack(pady=5, anchor='nw')

coin_label = tk.Label(app, text="Coins: 0", font=("Montserrat", 16), fg="gold")
coin_label.pack(pady=5, anchor='nw')

pause_icon = tk.PhotoImage(file='icons8-pause-48.png')
pauseBtn = tk.Button(app, image=pause_icon, command=pause)
pauseBtn.image = pause_icon
pauseBtn.pack(pady=10, padx=10, anchor='ne')

tk.Label(app, text="Quiz App", font=("Montserrat", 34), fg='#1581B4').pack(pady=12)

question_label = tk.Label(app, text="", font=('Montserrat', 16), wraplength=500, justify="center")
question_label.pack(pady=20)

button_frame = tk.Frame(app)
button_frame.pack(pady=10)

buttons = []
for i in range(4):
    btn = tk.Button(button_frame, text="", bg="#1581B4", fg="white", font=("Montserrat", 18), padx=10, pady=10, borderwidth=0, relief="flat", width=15)
    btn.grid(row=i//2, column=i % 2, padx=10, pady=10)
    buttons.append(btn)

prev_button = tk.Button(app, text="Previous", command=previous_question, font=("Montserrat", 14), bg="lightgray")
prev_button.pack(side=tk.LEFT)
next_button = tk.Button(app, text="Next", command=next_question, font=("Montserrat", 14), bg="lightgray")
next_button.pack(side=tk.LEFT)
submit_button = tk.Button(app, text="Submit", command=lambda: check_answer(None), font=("Montserrat", 14), bg="#1581B4", fg="black")
submit_button.pack(side=tk.LEFT)

question_section()
if questions:
    showquestion()

app.mainloop()
