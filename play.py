from tkinter import *
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import re
import random
import time
from tkinter import messagebox
from mainmenu import menuFunction
from store import storeCreator

root = Tk()
lightBlue = "#EDF7F7"
darkBlue = "#1581B4"

#get screen width and height
screenWidth = root.winfo_screenwidth()
screenHeight = root.winfo_screenheight()

#set the window size to match the screen dimensions
root.geometry(f"{screenWidth}x{screenHeight}")
root.title("Gameplay")
root.configure(bg=lightBlue)

playFrame = Frame(root, bg = lightBlue )
playFrame.place(x=0, y=0, relwidth=1, relheight= 1)

statusContain = Frame(playFrame, bg=lightBlue)
statusContain.pack(fill ="x", padx= 15, pady=15 )

#status bar
#statusBar = Label(statusContain, bg =lightBlue)
#statusBar.pack(side="left")

#image processing
saveMePic = Image.open("images/save me.png")
eliminatePic = Image.open("images/eliminate.png")
instantPic = Image.open("images/instant.png")
hintPic = Image.open("images/hint.png")
pausePic = Image.open("images/pause.png")
moneyPic = Image.open("images/money.png")

l1= Image.open("images/full life.png")
l2= Image.open("images/half life.png")
l3= Image.open("images/empty life.png")

powerUpImages = [ eliminatePic, instantPic, hintPic, pausePic, moneyPic] 
powerUpImageList =[]

#setting image height
for i in powerUpImages:
    myHeight = 40
    aspectRatio = i.width/i.height
    newWidth = int(myHeight * aspectRatio)

    resized = i.resize((newWidth, myHeight))
    tk_image = ImageTk.PhotoImage(resized)
    powerUpImageList.append(tk_image)

lifeImages = [l3,l2,l1]
lifeImageList = []

#setting image height
for i in lifeImages:
    myHeight = 25
    aspectRatio = i.width/i.height
    newWidth = int(myHeight * aspectRatio)

    resized = i.resize((newWidth, myHeight))
    tk_image = ImageTk.PhotoImage(resized)
    lifeImageList.append(tk_image)

lives = 2


#for the  life  meter
lifeMeter = Label(statusContain,
image=lifeImageList[2] ,
bg=lightBlue,
fg= darkBlue )
lifeMeter.pack(side="left")

def missed():
    global lifeMeter
    global lives
    print(lives)
    #close window if lives reaches 0
    if lives == 1:
        gameOver()

    lives = lives - 1
    lifeMeter.config(image = lifeImageList[lives])

workbook = load_workbook("users database.xlsx")
sheet = workbook.active
eliminateAmount = sheet["F2"].value
instantAmount = sheet["E2"].value
hintAmount = sheet["G2"].value

#money widget programming

#get money value
def getMoney():
    workbook = load_workbook("users database.xlsx")
    sheet = workbook.active
    return sheet["H2"].value


moneyValue = getMoney()
moneyContainer = Frame(statusContain, bg="#EDF7F7")
moneyContainer.pack(side="left", padx=20)

moneyIcon = Label(moneyContainer,
image = powerUpImageList[4],
bg ="#EDF7F7")
moneyIcon.pack(side="left", padx=5)

moneyAmount = Label(moneyContainer,
text= str(moneyValue),
bg="#EDF7F7", 
font= ("Montserrat", 14),
fg="#1581B4" )
moneyAmount.pack(side="left", padx=5)

#****************************************************************************************************************************
#timer stuff

timeLeft = 20
#to track if the timer is running
running = False

#store the ID to cancel it later
timerID = None

def updateTimer():
    """updates the countdown timer every second"""
    global timeLeft,  running, timerID
    if timeLeft > 0 and running:
        timeLeft -= 1
        timerLabel.config(text=f"Timer : {timeLeft}")
        timerID = root.after(1000, updateTimer)
    elif timeLeft == 0:
        timerLabel.config(text=f"Timer : {timeLeft}")
        missed()
        updateQuestions()

def startTimer():
    """starts the countdown"""
    global running, timerID

    #prevent multiple start calls
    if not running:
        running = True
        updateTimer()

def pauseTimer():
    """Pauses the countdown."""
    global running
    running = False
    if timerID:
        root.after_cancel(timerID)
    #stop any scheduled updates
    
def resetTimer():
    """resets the countdown"""
    global timeLeft, running, timerID
    running = False
    timeLeft = 20
    timerLabel.config(text=f"Timer : {timeLeft}")

    if timerID:
        root.after_cancel(timerID)
        timerID = None
        


#timer
timerLabel = Label(statusContain,
bg = lightBlue,
fg=darkBlue,
text= "Timer : 15",
font=("Montserrat", 14, "bold"))
timerLabel.pack(side="left", padx=50, anchor="center")

#******************************************************************************************************************************
#quitting theplay page
def onQuit(gameOverMsg):
    print("quit")
    #open mainmenu

    #don't change the variable name of this one
    #openMainMenu()
    gameOverMsg.destroy()

#how to access gameover from the pause page
def pauseGameOver(pauseMsg):
    gameOver()
    pauseMsg.destroy()

#game over message box
def gameOver():
    print("gameover")
    gameOverMsg = Toplevel()
    gameOverMsg.title("Pause")
    #setting the window size

    gameOverMsg.resizable(False, False)

    screenWidth = gameOverMsg.winfo_screenwidth()
    screenHeight = gameOverMsg.winfo_screenheight()


    #centering the window
    x = (screenWidth//2)-(300//2)
    y = (screenHeight//2)-(220//2)
    gameOverMsg.geometry(f"300x220+{x}+{y}")
    gameOverMsg.configure(bg=lightBlue)

    


    Label(gameOverMsg,bg=lightBlue).pack(pady=2)

    Label(gameOverMsg,
    text="Gameover",
    fg="black",
    bg=lightBlue,
    font=("Montserrat Extrabold", 17),
    justify="center").pack(pady=10)

    quitButton = Button(gameOverMsg,
    text="QUIT",
    bg=darkBlue,
    width=20,
    fg="white",
    command=lambda : onQuit(gameOverMsg),
    font=("Montserrat", 13, "bold"),
    pady=5).pack(pady=8)


    #make it appear on top of current window
    gameOverMsg.transient(root)

    #disable the main window
    gameOverMsg.grab_set()

    #wait for Pause window to close
    root.wait_window(gameOverMsg)




#********************************************************************************************************************************
#pause message box
def pauseResume(pauseMsg):
    startTimer()
    pauseMsg.destroy()

def pauseCommand():
    print("pause")
    pauseTimer()
    pauseMsg = Toplevel()
    pauseMsg.title("Pause")
    #setting the window size
    
    pauseMsg.resizable(False, False)

    screenWidth = pauseMsg.winfo_screenwidth()
    screenHeight = pauseMsg.winfo_screenheight()

    #centering the window
    x = (screenWidth//2)-(300//2)
    y = (screenHeight//2)-(220//2)
    pauseMsg.geometry(f"300x220+{x}+{y}")
    pauseMsg.configure(bg=lightBlue)

    Label(pauseMsg,bg=lightBlue).pack(pady=2)

    Label(pauseMsg,
    text="QuizLark",
    fg="black",
    bg=lightBlue,
    font=("Montserrat Extrabold", 16),
    justify="center").pack(pady=10)

    resumeButton = Button(pauseMsg,
    text="RESUME",
    bg=darkBlue,
    width=20,
    fg="white",
    command= lambda : pauseResume(pauseMsg),
    font=("Montserrat", 13, "bold"),
    pady=5).pack(pady=8)

    quitButton = Button(pauseMsg,
    text= "QUIT",
    bg=darkBlue,
    width=20,
    fg="white",
    font=("Montserrat", 13, "bold"),
    pady=5,
    command= lambda: pauseGameOver(pauseMsg)
    ).pack(
        pady=8
    )

    #make it appear on top of current window
    pauseMsg.transient(root)

    #disable the main window
    pauseMsg.grab_set()

    #wait for Pause window to close
    root.wait_window(pauseMsg)

    

#**************************************************************************************************************************
#pause button stuff
pauseContain = Frame(statusContain, bg = lightBlue)
pauseContain.pack(side="right", padx=40)

pauseButton = Button(pauseContain,
bg=lightBlue,
image=powerUpImageList[3],
command=pauseCommand,
relief="flat",
borderwidth=0
)
pauseButton.pack(side="left", padx=3)


#eliminate stuff
#*******************************************************************************************************************************
#*******************************************************************************************************************************

def eliminateCommand(): 
    global optionA
    global optionB
    global optionC
    global optionD
    global currentQuestion
    global currentAnswer
    global currentOptions 

    optionsList = []

    if optionA.cget("text") != currentAnswer:
        optionsList.append(optionA)

    if optionB.cget("text") != currentAnswer:
        optionsList.append(optionB)

    if optionC.cget("text") != currentAnswer:
        optionsList.append(optionC)

    if optionD.cget("text") != currentAnswer:
        optionsList.append(optionD)

    optionsList.pop(random.randint(0,2))
    for i in optionsList:
        i.grid_forget()

     #reduce the amount of hints in the database
    Workbook = load_workbook("users database.xlsx")
    sheet2 = Workbook.active
    amount = sheet2["F2"].value
    amount= amount - 1

    #change this
    eliminate.config(text=str(amount))

    #save changes
    workbook2 = load_workbook("users database.xlsx")
    sheet2 = workbook2.active
    sheet2["F2"] = amount
    workbook2.save("users database.xlsx")
    


eliminateContain = Frame(statusContain, bg = lightBlue)
eliminateContain.pack(side="right", padx=20)

eliminateButton = Button(eliminateContain,
bg=lightBlue,
image=powerUpImageList[0],
command=eliminateCommand,
relief="flat",
borderwidth=0
)
eliminateButton.pack(side="left", padx=3)

eliminate = Label(eliminateContain,
text=str(eliminateAmount),
bg ="#EDF7F7",
fg="black",
font=("Montserrat", 14, "bold"))
eliminate.pack(side="left")

#**************************************************************************************************************************
#Instant stuff
instantContain = Frame(statusContain, bg = lightBlue)
instantContain.pack(side="right", padx=20)

def instantCommand():
    global optionA
    global optionB
    global optionC
    global optionD
    global currentQuestion
    global currentAnswer
    global currentOptions 
    questionIndex = getIndex()
    text = optionA.cget("text")
    
    if optionA.cget("text") == currentAnswer:
        optionA.config(bg=darkBlue, fg="white")

    elif optionB.cget("text") == currentAnswer:
        optionB.config(bg=darkBlue, fg="white")

    elif optionC.cget("text") == currentAnswer:
        optionC.config(bg=darkBlue, fg="white")

    elif optionD.cget("text") == currentAnswer:
        optionD.config(bg=darkBlue, fg="white")

    #reduce the amount of hints in the database
    Workbook = load_workbook("users database.xlsx")
    sheet2 = Workbook.active
    amount = sheet2["E2"].value
    amount= amount - 1

    #change this
    instant.config(text=str(amount))

    #save changes
    workbook2 = load_workbook("users database.xlsx")
    sheet2 = workbook2.active
    sheet2["E2"] = amount
    workbook2.save("users database.xlsx")
    
    

instantButton = Button(instantContain,
bg=lightBlue,
image=powerUpImageList[1],
command=lambda : instantCommand(),
relief="flat",
borderwidth=0
)
instantButton.pack(side="left", padx=3)

instant = Label(instantContain,
text=str(instantAmount),
bg ="#EDF7F7",
fg="black",
font=("Montserrat", 14, "bold"))
instant.pack(side="left")


#************************************************************************************************************************
#**************************************************************************************************************************
#hint stuff
def hintCommand():
    global hintText
    global hint
    workbook = load_workbook("General_Knowledge_Quiz_50_Questions .xlsx")
    sheet = workbook.active
    #to access this question index variable that was created inside the get question function
    questionIndex = getIndex()
    currentHint = sheet[f"F{questionIndex}"].value
    hintText.config(text=str(currentHint))
    print(questionIndex)

    #reduce the amount of hints in the database
    Workbook = load_workbook("users database.xlsx")
    sheet2 = Workbook.active
    amount = sheet2["G2"].value
    amount= amount - 1

    #change this
    hint.config(text=str(amount))

    #save changes
    workbook2 = load_workbook("users database.xlsx")
    sheet2 = workbook2.active
    sheet2["G2"] = amount
    workbook2.save("users database.xlsx")


#hint contain
hintContain = Frame(statusContain, bg = lightBlue)
hintContain.pack(side="right", padx=20)



hintButton = Button(hintContain,
bg=lightBlue,
image=powerUpImageList[2],
command=hintCommand,
relief="flat",
borderwidth=0
)
hintButton.pack(side="left", padx=3)

hint = Label(hintContain,
text=str(hintAmount),
bg ="#EDF7F7",
fg="black",
font=("Montserrat", 14, "bold"))
hint.pack(side="left")
#****************************************************************************************************************************
#***************************************************************************************************************************

Label(playFrame,
bg=lightBlue).pack(pady=20)
#for the center frame
centerFrame = Frame(playFrame, bg="#EDF7F7")
centerFrame.pack()

Label(centerFrame,
text= "QuizLark",
font=("Montserrat Extrabold", 32),
justify="center",
bg=lightBlue,
fg="black").pack()

hintText = Label(centerFrame,
text = "",
font = ("Montserrat", 14),
fg="black",
bg = lightBlue)
hintText.pack()


questionsIndexes =  [x for x in range(1,51)]


#define function to get questions
def getQuestion():
    global questionsIndexes
    global questionIndex
    global getIndex

    def getIndex():
        return questionIndex

    questionIndex = random.choice(questionsIndexes)
    #pick a random index
    
    #remove the question chosen from the question list
    questionsIndexes.remove(questionIndex)

    #load questions database
    workbook = load_workbook("General_Knowledge_Quiz_50_Questions .xlsx")
    sheet = workbook.active

    question = sheet[f"A{questionIndex}"].value
    columns = ["B", "C", "D", "E"]

    #get the options and shuffle it
    options = [sheet[f"{column}{questionIndex}"].value for column in columns]
    random.shuffle(options)

    #get the correct answer
    correctAnswer = sheet[f"G{questionIndex}"].value

    return question, options, correctAnswer

def updateQuestions():
    global questionText
    global optionA
    global optionB
    global optionC
    global optionD
    global currentQuestion
    global currentAnswer
    global currentOptions
    currentQuestion, currentOptions, currentAnswer = getQuestion()
    #update changes
    questionText.config(text=currentQuestion)
    optionA.config(text=currentOptions[0], bg="white", fg="black")
    optionB.config(text=currentOptions[1], bg="white", fg="black")
    optionC.config(text=currentOptions[2], bg="white", fg="black")
    optionD.config(text=currentOptions[3], bg="white", fg="black")

    #putting the options on the page
    optionA.grid(row=0, column = 0, padx=10, pady=10)
    optionB.grid(row=0, column = 1, padx=10, pady=10)
    optionC.grid(row=1, column = 0, padx=10, pady=10)
    optionD.grid(row=1, column = 1, padx=10, pady=10)
    
    #restart the timer
    resetTimer()
    startTimer()

   


#when the user chooses an option
def questionAttempted(choice, answer, chosenOption):
    global moneyAmount
    if choice == answer:
        chosenOption.config(bg="green", fg="white")

        
        #add to money
        workbook = load_workbook("users database.xlsx")
        sheet = workbook.active
        previousAmount = sheet["H2"].value

        newAmount = previousAmount + 50
        moneyAmount.config(text=str(newAmount))

        workbook2 = load_workbook("users database.xlsx")
        sheet = workbook2.active
        sheet["H2"] = newAmount
        workbook2.save("users database.xlsx")


        root.after(1000, lambda : updateQuestions())
        
        
        
    elif choice != answer:
        chosenOption.config(bg="red", fg="white")
        root.after(1000, lambda : updateQuestions())
        missed()


#start the timer when question gets loaded
startTimer()

currentQuestion, currentOptions, currentAnswer = getQuestion()
questionFrame = Frame(centerFrame, 
width=400,
height = 150, 
bg="white",
pady=15,
padx=10)
questionFrame.pack(pady=15)

#defining the question text
questionText = Label(questionFrame,
text = str(currentQuestion),
font=("Montserrat", 18),
bg="white",
fg="black",
justify="left")
questionText.pack(padx=200, pady=40)

optionFrame = Frame(centerFrame, 
width=400,
height = 300, 
bg=lightBlue,
pady=15,
padx=10)
optionFrame.pack(pady=15)

optionA = Button(optionFrame,
text= str(currentOptions[0]),
bg = "white",
width= 30,
pady=10,
fg= "black",
font=("Montserrat", 16),
justify="left",
command= lambda : questionAttempted(currentOptions[0], currentAnswer, optionA))

optionB = Button(optionFrame,
text= str(currentOptions[1]),
bg = "white",
width= 30,
pady = 10,
fg= "black",
font=("Montserrat", 16),
justify="left",
command= lambda : questionAttempted(currentOptions[1], currentAnswer, optionB))

optionC = Button(optionFrame,
text= str(currentOptions[2]),
bg = "white",
width= 30,
pady=10,
fg= "black",
font=("Montserrat", 16),
justify="left",
command= lambda : questionAttempted(currentOptions[2], currentAnswer, optionC))

optionD = Button(optionFrame,
text= str(currentOptions[3]),
width= 30,
pady= 10,
bg = "white",
fg= "black",
font=("Montserrat", 16),
justify="left",
command= lambda : questionAttempted(currentOptions[3], currentAnswer, optionD))

#putting the options on the page
optionA.grid(row=0, column = 0, padx=10, pady=10)
optionB.grid(row=0, column = 1, padx=10, pady=10)
optionC.grid(row=1, column = 0, padx=10, pady=10)
optionD.grid(row=1, column = 1, padx=10, pady=10)




root.mainloop()