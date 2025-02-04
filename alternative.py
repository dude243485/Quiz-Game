#note that this color: "#1581B4" is the light blue use in the app
#while this color: "#EDF7F7" is the dark blue used in the app



from tkinter import *
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import re
import random
from tkinter import messagebox
from mainmenu import menuFunction
from store import storeCreator


lightBlue = "#EDF7F7"
darkBlue = "#1581B4"
userIndex = 2

#creating the root window
root = Tk()
root.title("QuizLark App")

#get screen width and height
screenWidth = root.winfo_screenwidth()
screenHeight = root.winfo_screenheight()

#set the window size to match the screen dimensions
root.geometry(f"{screenWidth}x{screenHeight}")

#*****************************************************************************************************************************
#welcome page functions


#this function changes the welcome page to the login page
#by destorying frames
def loginCommand():
    #globalize variables to allow outside access
    global loginFrame
    global welcomeFrame

    #close welcome frame
    loginFrame.tkraise()
    
    #displaying the current login frame
    loginFrame.place(x=0, y=0, relwidth=1, relheight= 1)

def update(amountInStock, index):
    #load current value
    if index == 0:
        Workbook = load_workbook("users database.xlsx")
        sheet = Workbook.active
        save = sheet[f"D{userIndex}"].value
        save = save + 1
        amountInStock[0].configure(text=f"In-stock : {save}")

        #save changes
        workbook2 = load_workbook("users database.xlsx")
        sheet = workbook2.active
        sheet[f"D{userIndex}"] = save
        workbook2.save("users database.xlsx")

    elif index == 1:
        Workbook = load_workbook("users database.xlsx")
        sheet = Workbook.active
        amount = sheet[f"E{userIndex}"].value
        amount= amount + 1
        amountInStock[1].configure(text=f"In-stock : {amount}")

        #save changes
        workbook2 = load_workbook("users database.xlsx")
        sheet = workbook2.active
        sheet[f"E{userIndex}"] = amount
        workbook2.save("users database.xlsx")

    elif index == 2:
        Workbook = load_workbook("users database.xlsx")
        sheet = Workbook.active
        amount = sheet[f"F{userIndex}"].value
        amount= amount + 1
        amountInStock[2].configure(text=f"In-stock : {amount}")

        #save changes
        workbook2 = load_workbook("users database.xlsx")
        sheet = workbook2.active
        sheet[f"F{userIndex}"] = amount
        workbook2.save("users database.xlsx")
    elif index == 3:
        Workbook = load_workbook("users database.xlsx")
        sheet = Workbook.active
        amount = sheet[f"G{userIndex}"].value
        amount= amount + 1
        amountInStock[3].configure(text=f"In-stock : {amount}")

        #save changes
        workbook2 = load_workbook("users database.xlsx")
        sheet = workbook2.active
        sheet[f"G{userIndex}"] = amount
        workbook2.save("users database.xlsx")



def storeCommand():
    #get the store button and mainmenu frame from the menu function
    mainMenuFrame, storeButton, playButton = menuFunction(root, userIndex)
    myStore,amountInStock,buyButtons = storeCreator(root,mainMenuFrame, userIndex)

    #raise the store 
    myStore.tkraise()

    #assing the command for the store button
    storeButton.config(command = storeCommand)

    #assign the  command for the buy buttons and label the buttons
    buyButtons[0].config(command= lambda : update(amountInStock, 0))
    buyButtons[1].config(command = lambda : update(amountInStock, 1))
    buyButtons[2].config(command = lambda : update(amountInStock, 2))
    buyButtons[3].config(command = lambda : update(amountInStock, 3))

    
 #*******************************************************************************************************************************   
def openPlay():
    global lives, lifeMeter, timeLeft, running, timerID
    
    timeLeft = 20
    #to track if the timer is running
    running = False

    #store the ID to cancel it later
    timerID = None

    startTimer()

    lifeMeter.config(image=lifeImageList[2] )
    lives = 2
    playFrame.place(x=0, y=0, relwidth=1, relheight= 1)
    playFrame.tkraise()
    
def openMainMenu():
    mainMenuFrame, storeButton, playButton = menuFunction(root, userIndex)
    myStore,amountInStock,buyButtons = storeCreator(root, mainMenuFrame, userIndex)
    #assing the command for the store button
    storeButton.config(command = storeCommand)
    playButton.config(command = openPlay)

    #assign the  command for the buy buttons and label the buttons
    buyButtons[0].config(command= lambda : update(amountInStock, 0))
    buyButtons[1].config(command = lambda : update(amountInStock, 1))
    buyButtons[2].config(command = lambda : update(amountInStock, 2))
    buyButtons[3].config(command = lambda : update(amountInStock, 3))

    mainMenuFrame.tkraise()
    mainMenuFrame.place(x=0, y=0, relwidth=1, relheight= 1)


def accountCommand():
    global accountFrame
    #raise the account frame
    accountFrame.tkraise()
    accountFrame.place(x=0, y=0, relwidth=1, relheight=1)

def playAsGuestCommand():
    mainMenuFrame, storeButton, playButton = menuFunction(root, userIndex)
    storeButton.config(command=storeCommand)
    playButton.config(command = openPlay)
    mainMenuFrame.tkraise()
    mainMenuFrame.place(x=0, y=0, relwidth=1, relheight= 1)


    

#this is the frame that holds everything on our welcome page
welcomeFrame = Frame(root, bg="#EDF7F7")
welcomeFrame.place(x=0, y=0, relwidth=1, relheight= 1)


#open the image and convert it to tkinter format
backgroundImage = Image.open("images/background image.jpg")
tk_image = ImageTk.PhotoImage(backgroundImage)

#place the background image on the page
#backgroundLabel = Label(root, bg= "#EDF7F7")
#backgroundLabel.place(x=0, y=0, relwidth=1, relheight= 1)

root.grid_rowconfigure(0,weight=1)
root.grid_columnconfigure(0, weight=1)

#extra space on top
spaceLabel = Label(welcomeFrame).pack(pady=40)



#company name
welcomeLabel = Label(welcomeFrame,
text= "welcome", 
fg= "#1581B4", 
bg = "#EDF7F7",
font=("Montserrat Extrabold",18, "bold"), 
pady=1).pack(pady=0)

brandName = Label(welcomeFrame,
text= "QuizLark", 
fg= "black", 
bg = "#EDF7F7", 
font=("Montserrat Extrabold",36, "bold" ), pady=1).pack(pady=0)

#create the buttons container
myButtons = Label(welcomeFrame, 
bg = "#EDF7F7")
myButtons.place(relx=0.5, rely=0.5,  anchor= CENTER)

#creating the homepage buttons
playAsGuest = Button(myButtons, 
text="Play As Guest",
fg = "white",
bg="#1581B4", 
font=("Montserrat Extrabold", 16, "bold"),
padx=30,
pady=3,
command=playAsGuestCommand)

createAccount = Button(myButtons,
text="Create Account",
fg="white",
bg = "#1581B4", 
font=("Montserrat Extrabold", 16, "bold"), 
padx=30,
pady=3,
command = accountCommand)

login = Button(myButtons,
text="Log In",
fg="white",
bg = "#1581B4",
font=("Montserrat Extrabold", 16, "bold"), 
padx=30,
pady=3,
command= loginCommand)

#placing the buttons on the page
playAsGuest.grid(row = 0, column = 1, pady=3, padx=3)
createAccount.grid(row= 0, column=2, pady=3, padx=20)
login.grid(row=0, column = 3, pady=3, padx=3)


#******************************************************************************************************************************
#login page
#regex password creation [A-Za-z0-9@#!%+=]
def checkPasswordStrength(s):
    upperCase = bool(re.search(r'[A-Z]', s))
    lowerCase = bool(re.search(r'[a-z]', s))
    digits = bool(re.search(r"\d", s))
    special = bool(re.search(r"[^A-Za-z0-9]", s))

    if upperCase and lowerCase and digits and special:
        return True

    return False

#to check if users password is in password database
def checkCorrectPassword(s):
    workbook = load_workbook("users database.xlsx")
    sheet = workbook["Sheet1"]
    #loop through each row in the password column
    for cell in sheet["B"]:
        if s == cell.value:
            return True
        
    return False

#to check if user's username is in the username database
def checkUsername(n):
    workbook = load_workbook("users database.xlsx")
    sheet = workbook["Sheet1"]
    #loop through each row in the username column
    for cell in sheet["A"]:
        if n == cell.value:
            return True

    return False


#check if username and password match
def checkCorrectDetails(s,n):
    workbook = load_workbook("users database.xlsx")
    sheet = workbook.active

    userNameRowNumber = None
    #iterating through each cell to look for current password
    for row in sheet.iter_rows(min_row=1, max_row = sheet.max_row, min_col=1, max_col = sheet.max_column):
        #looking through each row
        for cell in row:
            if cell.value == n:
                userNameRowNumber = cell.row
                break
        #if this variable has been assigned a value then break
        if userNameRowNumber:
            break
def findRow(username):
    workbook = load_workbook("users database.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, max_row = sheet.max_row, min_col=1, max_col = sheet.max_column):
        for cell in row:
            if cell.value == username:
                return cell.row
    return None

    #for password row number
    passwordRowNumber = None
    for row in sheet.iter_rows(min_row=1, max_row = sheet.max_row, min_col=1, max_col = sheet.max_column):
        for cell in row:
            if cell.value == n:
                passwordRowNumber = cell.row
                break
        #if this variable has been assigned a value then break
        if passwordRowNumber:
            break

    #finally check is the password and username have the same row number
    if passwordRowNumber == userNameRowNumber:
        return True

    return False


    

#login button clicked function
def onLogin():
    global userIndex

    #globalize and store entered password
    global password
    password = passwordEntry.get()
    username = usernameEntry.get()


    #if correct password load mainmenu page
    if checkUsername(username) and checkCorrectPassword(password):
        if checkCorrectDetails(password, username):
            print("mainMenu Accessed")
            #open main menu
            openMainMenu()


    else:
        messagebox.showwarning("Weak Password", 
        "Password must contain uppercase, lowercase, special characters and digits!")
        print("Password or username is incorrect")
    userIndex = findRow(username)

            

#login frame creation
loginFrame = Frame(root, bg="#EDF7F7")
formContainer = Label(loginFrame,
bg="white",
width=450,
height=500)
formContainer.place(relx=0.5, rely=0.5,  anchor= CENTER)

#extra space on top
Label(formContainer, bg="white").pack(pady=10)

#company name
loginBrandName = Label(formContainer,
text= "QuizLark",
fg= "#1581B4",
bg = "white",
font=("Montserrat Extrabold",24, "bold"),
pady=1).pack(pady=0)

brandName = Label(formContainer,
text= "Login to your account",
fg= "black",bg = "white", 
font=("Montserrat Extrabold",16, "bold" ),
pady=1).pack(pady=0)

#extra space in between
Label(formContainer, bg="white").pack(pady=10)

#username label and entry
usernameLabel= Label(formContainer,
text="username",
font= ("Montserrat", 14),
bg="white").pack(anchor=W, padx=40)

usernameEntry = Entry(formContainer, 
width= 30,
bg="#EDF7F7",
font=("Montserrat, 14"))
usernameEntry.pack(anchor = W, padx=40, pady=10)

#password label and entry
passwordLabel= Label(formContainer,
text="password",
font= ("Montserrat", 14),
bg="white").pack(anchor=W, padx=40)

passwordEntry = Entry(formContainer,
width= 30,
bg="#EDF7F7",
font=("Montserrat, 14"))
passwordEntry.pack(anchor = W, padx=40, pady=10)



#extra space in between
Label(formContainer, bg="white" ).pack(pady=4)

#login button
loginButton = Button(formContainer, 
text="Login",
bg="#1581B4",
pady=8,
width=23,  
fg="white",
font=("Montserrat", 16, "bold"),
command = onLogin).pack( pady=10, padx=40)

#extra space in between
Label(formContainer, bg="white" ).pack(pady=5)

#*******************************************************************************************************************************
#create account page

#if user clicked create account
def onCreateAccount():
    global userIndex

    name = fullnameEntry.get()
    accountPassword = accountPasswordEntry.get()
    accountUsername = accountUsernameEntry.get()

    userIndex = findRow(accountUsername)

    #show error if username already exists
    if checkUsername(accountUsername):
        messagebox.showwarning("Error", "Username already exists")

    #show error message if password is weak
    if not checkPasswordStrength(accountPassword):
        messagebox.showwarning("Weak Password", 
        "Password must contain uppercase, lowercase, special characters and digits!")

    else:
        userData = [accountUsername, accountPassword, name]

        #writing current user data to the external database
        Dworkbook = load_workbook("users database.xlsx")
        sheet = Dworkbook.active
        sheet.append(userData)

        #save changes
        Dworkbook.save("users database.xlsx")
        print("Main menu Accessed")

        #open main menu
        openMainMenu()

#login frame creation

accountFrame = Frame(root, bg="#EDF7F7")
accountFormContainer = Label(accountFrame, 
bg="white",
width=600,
height=500)
accountFormContainer.pack(side="right", fill="y", padx=120, pady=120)

#create account greeting
greetingContainer = Label(accountFrame,
bg="#EDF7F7",
width= 450,
height=500,
padx=50,
pady=50)
greetingContainer.pack(side="left", fill="y", pady=120, padx=110)

Label(greetingContainer, 
text="QuizLark",
font= ("Montserrat", 36, "bold"),
fg="#1581B4",
bg="#EDF7F7").pack(anchor=W)

Label(greetingContainer,
text="Welcome to the best quizzing app out there...",
wraplength=380,
font=("Montserrat", 18, "bold"),
fg="black",
bg="#EDF7F7",
justify="left").pack(anchor=W,padx=10)

#extra space on top
Label(accountFormContainer, bg="white").pack(pady=10)

#company name
loginBrandName = Label(accountFormContainer,
text= "QuizLark",
fg= "#1581B4",
bg = "white",
font=("Montserrat Extrabold",24, "bold"),
pady=1).pack(pady=0)

brandName = Label(accountFormContainer,
text= "Create an account with us",
fg= "black",
bg = "white", 
font=("Montserrat Extrabold",16, "bold" ),
pady=1).pack(pady=0)

#extra space in between
Label(accountFormContainer, 
bg="white").pack(pady=10)

#username label and entry
fullnameLabel= Label(accountFormContainer,
text="Fullname",
font= ("Montserrat", 14),
bg="white").pack(anchor=W, padx=40)

fullnameEntry = Entry(accountFormContainer,
width= 30,
bg="#EDF7F7",
font=("Montserrat, 14"))
fullnameEntry.pack(anchor = W, padx=40, pady=10)


#username label and entry
accountUsernameLabel= Label(accountFormContainer,
text="username",
font= ("Montserrat", 14),
bg="white").pack(anchor=W, padx=40)

accountUsernameEntry = Entry(accountFormContainer,
width= 30,
bg="#EDF7F7",
font=("Montserrat, 14"))
accountUsernameEntry.pack(anchor = W, padx=40, pady=10)

#password label and entry
accountPasswordLabel= Label(accountFormContainer, 
text="password",
font= ("Montserrat", 14),
bg="white").pack(anchor=W, padx=40)

accountPasswordEntry = Entry(accountFormContainer,
width= 30,
bg="#EDF7F7",
font=("Montserrat, 14"))
accountPasswordEntry.pack(anchor = W, padx=40, pady=10)


#login button
createAccount = Button(accountFormContainer, 
text="Create Account",
bg="#1581B4",
pady=8, width=23,  
fg="white", font=("Montserrat", 16, "bold"),
command = onCreateAccount).pack( pady=10, padx=40)

#extra space in between
Label(accountFormContainer, bg="white" ).pack(pady=5)
#*****************************************************************************************************************************
#****************************************************************************************************************************
playFrame = Frame(root, bg = lightBlue )


statusContain = Frame(playFrame, bg=lightBlue)
statusContain.pack(fill ="x", padx= 15, pady=15 )

#status bar
#statusBar = Label(statusContain, bg =lightBlue)
#statusBar.pack(side="left")

#image processing
saveMePic = Image.open("images/save me.png")
eliminatePic = Image.open("images/eliminate.png")
instantPic = Image.open("images/instant.png")
hintPic = Image.open("images/hint.png")
pausePic = Image.open("images/pause.png")
moneyPic = Image.open("images/money.png")

l1= Image.open("images/full life.png")
l2= Image.open("images/half life.png")
l3= Image.open("images/empty life.png")

powerUpImages = [ eliminatePic, instantPic, hintPic, pausePic, moneyPic] 
powerUpImageList =[]

#setting image height
for i in powerUpImages:
    myHeight = 40
    aspectRatio = i.width/i.height
    newWidth = int(myHeight * aspectRatio)

    resized = i.resize((newWidth, myHeight))
    tk_image = ImageTk.PhotoImage(resized)
    powerUpImageList.append(tk_image)

lifeImages = [l3,l2,l1]
lifeImageList = []

#setting image height
for i in lifeImages:
    myHeight = 25
    aspectRatio = i.width/i.height
    newWidth = int(myHeight * aspectRatio)

    resized = i.resize((newWidth, myHeight))
    tk_image = ImageTk.PhotoImage(resized)
    lifeImageList.append(tk_image)

lives = 2


#for the  life  meter
lifeMeter = Label(statusContain,
image=lifeImageList[2] ,
bg=lightBlue,
fg= darkBlue )
lifeMeter.pack(side="left")

def missed():
    global lifeMeter
    global lives
    print(lives)
    #close window if lives reaches 0
    if lives == 1:
        pauseTimer()
        gameOver()

    lives = lives - 1
    lifeMeter.config(image = lifeImageList[lives])

workbook = load_workbook("users database.xlsx")
sheet = workbook.active
eliminateAmount = sheet[f"F{userIndex}"].value
instantAmount = sheet[f"E{userIndex}"].value
hintAmount = sheet[f"G{userIndex}"].value

#money widget programming

#get money value
def getMoney():
    workbook = load_workbook("users database.xlsx")
    sheet = workbook.active
    return sheet[f"H{userIndex}"].value


moneyValue = getMoney()
moneyContainer = Frame(statusContain, bg="#EDF7F7")
moneyContainer.pack(side="left", padx=20)

moneyIcon = Label(moneyContainer,
image = powerUpImageList[4],
bg ="#EDF7F7")
moneyIcon.pack(side="left", padx=5)

moneyAmount = Label(moneyContainer,
text= str(moneyValue),
bg="#EDF7F7", 
font= ("Montserrat", 14),
fg="#1581B4" )
moneyAmount.pack(side="left", padx=5)

#****************************************************************************************************************************
#timer stuff

timeLeft = 20
#to track if the timer is running
running = False

#store the ID to cancel it later
timerID = None

def updateTimer():
    """updates the countdown timer every second"""
    global timeLeft,  running, timerID
    if timeLeft > 0 and running:
        timeLeft -= 1
        timerLabel.config(text=f"Timer : {timeLeft}")
        timerID = root.after(1000, updateTimer)
    elif timeLeft == 0:
        timerLabel.config(text=f"Timer : {timeLeft}")
        missed()
        updateQuestions()

def startTimer():
    """starts the countdown"""
    global running, timerID

    #prevent multiple start calls
    if not running:
        running = True
        updateTimer()

def pauseTimer():
    """Pauses the countdown."""
    global running
    running = False
    if timerID:
        root.after_cancel(timerID)
    #stop any scheduled updates
    
def resetTimer():
    """resets the countdown"""
    global timeLeft, running, timerID
    running = False
    timeLeft = 20
    timerLabel.config(text=f"Timer : {timeLeft}")

    if timerID:
        root.after_cancel(timerID)
        timerID = None
        


#timer
timerLabel = Label(statusContain,
bg = lightBlue,
fg=darkBlue,
text= "Timer : 15",
font=("Montserrat", 14, "bold"))
timerLabel.pack(side="left", padx=50, anchor="center")

#******************************************************************************************************************************
#quitting theplay page
def onQuit(gameOverMsg):
    #open mainmenu
    pauseTimer()

    #don't change the variable name of this one
    openMainMenu()
    gameOverMsg.destroy()

#how to access gameover from the pause page
def pauseGameOver(pauseMsg):
    pauseTimer()
    gameOver()
    pauseMsg.destroy()

#game over message box
def gameOver():
    print("gameover")
    
    gameOverMsg = Toplevel()
    gameOverMsg.title("Pause")
    #setting the window size

    gameOverMsg.resizable(False, False)

    screenWidth = gameOverMsg.winfo_screenwidth()
    screenHeight = gameOverMsg.winfo_screenheight()


    #centering the window
    x = (screenWidth//2)-(300//2)
    y = (screenHeight//2)-(220//2)
    gameOverMsg.geometry(f"300x220+{x}+{y}")
    gameOverMsg.configure(bg=lightBlue)

    


    Label(gameOverMsg,bg=lightBlue).pack(pady=2)

    Label(gameOverMsg,
    text="Gameover",
    fg="black",
    bg=lightBlue,
    font=("Montserrat Extrabold", 17),
    justify="center").pack(pady=10)

    quitButton = Button(gameOverMsg,
    text="QUIT",
    bg=darkBlue,
    width=20,
    fg="white",
    command=lambda : onQuit(gameOverMsg),
    font=("Montserrat", 13, "bold"),
    pady=5).pack(pady=8)


    #make it appear on top of current window
    gameOverMsg.transient(root)

    #disable the main window
    gameOverMsg.grab_set()

    #wait for Pause window to close
    root.wait_window(gameOverMsg)




#********************************************************************************************************************************
#pause message box
def pauseResume(pauseMsg):
    startTimer()
    pauseMsg.destroy()

def pauseCommand():
    print("pause")
    pauseTimer()
    pauseMsg = Toplevel()
    pauseMsg.title("Pause")
    #setting the window size
    
    pauseMsg.resizable(False, False)

    screenWidth = pauseMsg.winfo_screenwidth()
    screenHeight = pauseMsg.winfo_screenheight()

    #centering the window
    x = (screenWidth//2)-(300//2)
    y = (screenHeight//2)-(220//2)
    pauseMsg.geometry(f"300x220+{x}+{y}")
    pauseMsg.configure(bg=lightBlue)

    Label(pauseMsg,bg=lightBlue).pack(pady=2)

    Label(pauseMsg,
    text="QuizLark",
    fg="black",
    bg=lightBlue,
    font=("Montserrat Extrabold", 16),
    justify="center").pack(pady=10)

    resumeButton = Button(pauseMsg,
    text="RESUME",
    bg=darkBlue,
    width=20,
    fg="white",
    command= lambda : pauseResume(pauseMsg),
    font=("Montserrat", 13, "bold"),
    pady=5).pack(pady=8)

    quitButton = Button(pauseMsg,
    text= "QUIT",
    bg=darkBlue,
    width=20,
    fg="white",
    font=("Montserrat", 13, "bold"),
    pady=5,
    command= lambda: pauseGameOver(pauseMsg)
    ).pack(
        pady=8
    )

    #make it appear on top of current window
    pauseMsg.transient(root)

    #disable the main window
    pauseMsg.grab_set()

    #wait for Pause window to close
    root.wait_window(pauseMsg)

    

#**************************************************************************************************************************
#pause button stuff
pauseContain = Frame(statusContain, bg = lightBlue)
pauseContain.pack(side="right", padx=40)

pauseButton = Button(pauseContain,
bg=lightBlue,
image=powerUpImageList[3],
command=pauseCommand,
relief="flat",
borderwidth=0
)
pauseButton.pack(side="left", padx=3)


#eliminate stuff
#*******************************************************************************************************************************
#*******************************************************************************************************************************

def eliminateCommand(): 
    global optionA
    global optionB
    global optionC
    global optionD
    global currentQuestion
    global currentAnswer
    global currentOptions 

    optionsList = []

    if optionA.cget("text") != currentAnswer:
        optionsList.append(optionA)

    if optionB.cget("text") != currentAnswer:
        optionsList.append(optionB)

    if optionC.cget("text") != currentAnswer:
        optionsList.append(optionC)

    if optionD.cget("text") != currentAnswer:
        optionsList.append(optionD)

    optionsList.pop(random.randint(0,2))
    for i in optionsList:
        i.grid_forget()

     #reduce the amount of hints in the database
    Workbook = load_workbook("users database.xlsx")
    sheet2 = Workbook.active
    amount = sheet2[f"F{userIndex}"].value
    amount= amount - 1

    #change this
    eliminate.config(text=str(amount))

    #save changes
    workbook2 = load_workbook("users database.xlsx")
    sheet2 = workbook2.active
    sheet2[f"F{userIndex}"] = amount
    workbook2.save("users database.xlsx")
    


eliminateContain = Frame(statusContain, bg = lightBlue)
eliminateContain.pack(side="right", padx=20)

eliminateButton = Button(eliminateContain,
bg=lightBlue,
image=powerUpImageList[0],
command=eliminateCommand,
relief="flat",
borderwidth=0
)
eliminateButton.pack(side="left", padx=3)

eliminate = Label(eliminateContain,
text=str(eliminateAmount),
bg ="#EDF7F7",
fg="black",
font=("Montserrat", 14, "bold"))
eliminate.pack(side="left")

#**************************************************************************************************************************
#Instant stuff
instantContain = Frame(statusContain, bg = lightBlue)
instantContain.pack(side="right", padx=20)

def instantCommand():
    global optionA
    global optionB
    global optionC
    global optionD
    global currentQuestion
    global currentAnswer
    global currentOptions 
    questionIndex = getIndex()
    text = optionA.cget("text")
    
    if optionA.cget("text") == currentAnswer:
        optionA.config(bg=darkBlue, fg="white")

    elif optionB.cget("text") == currentAnswer:
        optionB.config(bg=darkBlue, fg="white")

    elif optionC.cget("text") == currentAnswer:
        optionC.config(bg=darkBlue, fg="white")

    elif optionD.cget("text") == currentAnswer:
        optionD.config(bg=darkBlue, fg="white")

    #reduce the amount of hints in the database
    Workbook = load_workbook("users database.xlsx")
    sheet2 = Workbook.active
    amount = sheet2[f"E{userIndex}"].value
    amount= amount - 1

    #change this
    instant.config(text=str(amount))

    #save changes
    workbook2 = load_workbook("users database.xlsx")
    sheet2 = workbook2.active
    sheet2[f"E{userIndex}"] = amount
    workbook2.save("users database.xlsx")
    
    

instantButton = Button(instantContain,
bg=lightBlue,
image=powerUpImageList[1],
command=lambda : instantCommand(),
relief="flat",
borderwidth=0
)
instantButton.pack(side="left", padx=3)

instant = Label(instantContain,
text=str(instantAmount),
bg ="#EDF7F7",
fg="black",
font=("Montserrat", 14, "bold"))
instant.pack(side="left")


#************************************************************************************************************************
#**************************************************************************************************************************
#hint stuff
def hintCommand():
    global hintText
    global hint
    workbook = load_workbook("General_Knowledge_Quiz_50_Questions .xlsx")
    sheet = workbook.active
    #to access this question index variable that was created inside the get question function
    questionIndex = getIndex()
    currentHint = sheet[f"F{questionIndex}"].value
    hintText.config(text=str(currentHint))
    print(questionIndex)

    #reduce the amount of hints in the database
    Workbook = load_workbook("users database.xlsx")
    sheet2 = Workbook.active
    amount = sheet2[f"G{userIndex}"].value
    amount= amount - 1

    #change this
    hint.config(text=str(amount))

    #save changes
    workbook2 = load_workbook("users database.xlsx")
    sheet2 = workbook2.active
    sheet2[f"G{userIndex}"] = amount
    workbook2.save("users database.xlsx")


#hint contain
hintContain = Frame(statusContain, bg = lightBlue)
hintContain.pack(side="right", padx=20)



hintButton = Button(hintContain,
bg=lightBlue,
image=powerUpImageList[2],
command=hintCommand,
relief="flat",
borderwidth=0
)
hintButton.pack(side="left", padx=3)

hint = Label(hintContain,
text=str(hintAmount),
bg ="#EDF7F7",
fg="black",
font=("Montserrat", 14, "bold"))
hint.pack(side="left")
#****************************************************************************************************************************
#***************************************************************************************************************************

Label(playFrame,
bg=lightBlue).pack(pady=20)
#for the center frame
centerFrame = Frame(playFrame, bg="#EDF7F7")
centerFrame.pack()

Label(centerFrame,
text= "QuizLark",
font=("Montserrat Extrabold", 32),
justify="center",
bg=lightBlue,
fg="black").pack()

hintText = Label(centerFrame,
text = "",
font = ("Montserrat", 14),
fg="black",
bg = lightBlue)
hintText.pack()


questionsIndexes =  [x for x in range(1,51)]


#define function to get questions
def getQuestion():
    global questionsIndexes
    global questionIndex
    global getIndex

    def getIndex():
        return questionIndex

    questionIndex = random.choice(questionsIndexes)
    #pick a random index
    
    #remove the question chosen from the question list
    questionsIndexes.remove(questionIndex)

    #load questions database
    workbook = load_workbook("General_Knowledge_Quiz_50_Questions .xlsx")
    sheet = workbook.active

    question = sheet[f"A{questionIndex}"].value
    columns = ["B", "C", "D", "E"]

    #get the options and shuffle it
    options = [sheet[f"{column}{questionIndex}"].value for column in columns]
    random.shuffle(options)

    #get the correct answer
    correctAnswer = sheet[f"G{questionIndex}"].value

    return question, options, correctAnswer

def updateQuestions():
    global questionText
    global optionA
    global optionB
    global optionC
    global optionD
    global currentQuestion
    global currentAnswer
    global currentOptions
    global hintText
    
    currentQuestion, currentOptions, currentAnswer = getQuestion()
    #update changes
    questionText.config(text=currentQuestion)
    optionA.config(text=currentOptions[0], bg="white", fg="black")
    optionB.config(text=currentOptions[1], bg="white", fg="black")
    optionC.config(text=currentOptions[2], bg="white", fg="black")
    optionD.config(text=currentOptions[3], bg="white", fg="black")

    #putting the options on the page
    optionA.grid(row=0, column = 0, padx=10, pady=10)
    optionB.grid(row=0, column = 1, padx=10, pady=10)
    optionC.grid(row=1, column = 0, padx=10, pady=10)
    optionD.grid(row=1, column = 1, padx=10, pady=10)
    hintText.config(text="")
    
    if lives >= 1:
        #restart the timer
        resetTimer()
        startTimer()

   


#when the user chooses an option
def questionAttempted(choice, answer, chosenOption):
    global moneyAmount
    if choice == answer:
        chosenOption.config(bg="green", fg="white")

        
        #add to money
        workbook = load_workbook("users database.xlsx")
        sheet = workbook.active
        previousAmount = sheet[f"H{userIndex}"].value

        newAmount = previousAmount + 50
        moneyAmount.config(text=str(newAmount))

        workbook2 = load_workbook("users database.xlsx")
        sheet = workbook2.active
        sheet[f"H{userIndex}"] = newAmount
        workbook2.save("users database.xlsx")


        root.after(1000, lambda : updateQuestions())
        
        
        
    elif choice != answer:
        chosenOption.config(bg="red", fg="white")
        root.after(1000, lambda : updateQuestions())
        missed()



currentQuestion, currentOptions, currentAnswer = getQuestion()
questionFrame = Frame(centerFrame, 
width=400,
height = 150, 
bg="white",
pady=15,
padx=10)
questionFrame.pack(pady=15)

#defining the question text
questionText = Label(questionFrame,
text = str(currentQuestion),
font=("Montserrat", 18),
bg="white",
fg="black",
justify="left")
questionText.pack(padx=200, pady=40)

optionFrame = Frame(centerFrame, 
width=400,
height = 300, 
bg=lightBlue,
pady=15,
padx=10)
optionFrame.pack(pady=15)

optionA = Button(optionFrame,
text= str(currentOptions[0]),
bg = "white",
width= 30,
pady=10,
fg= "black",
font=("Montserrat", 16),
justify="left",
command= lambda : questionAttempted(currentOptions[0], currentAnswer, optionA))

optionB = Button(optionFrame,
text= str(currentOptions[1]),
bg = "white",
width= 30,
pady = 10,
fg= "black",
font=("Montserrat", 16),
justify="left",
command= lambda : questionAttempted(currentOptions[1], currentAnswer, optionB))

optionC = Button(optionFrame,
text= str(currentOptions[2]),
bg = "white",
width= 30,
pady=10,
fg= "black",
font=("Montserrat", 16),
justify="left",
command= lambda : questionAttempted(currentOptions[2], currentAnswer, optionC))

optionD = Button(optionFrame,
text= str(currentOptions[3]),
width= 30,
pady= 10,
bg = "white",
fg= "black",
font=("Montserrat", 16),
justify="left",
command= lambda : questionAttempted(currentOptions[3], currentAnswer, optionD))

#putting the options on the page
optionA.grid(row=0, column = 0, padx=10, pady=10)
optionB.grid(row=0, column = 1, padx=10, pady=10)
optionC.grid(row=1, column = 0, padx=10, pady=10)
optionD.grid(row=1, column = 1, padx=10, pady=10)

root.mainloop()

