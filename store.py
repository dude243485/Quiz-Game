from tkinter import *
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import re
from tkinter import messagebox
from functools import partial


        

def storeCreator(root, mainMenuFrame, userIndex):


    global saveMePic
    global eliminatePic
    global instantPic
    global hintPic 
    global images 
    global imageList 
    global backButton
   

    #your functions
    def goBack():
        mainMenuFrame.tkraise()



    #create the store frame
    storeFrame = Frame(root, bg="#EDF7F7")
    storeFrame.place(x=0, y=0, relwidth=1, relheight= 1)

    #open the image
    backButton = Image.open("images/back arrow.png")

    #define your chosen height
    backButtonHeight = 50

    #get the image aspect ratio
    backButtonRatio = backButton.width/backButton.height

    #use the aspect ratio to calculate the new width
    newWidth = int(backButtonHeight*backButtonRatio)

    #resize the image
    resized = backButton.resize((newWidth, backButtonHeight))
    backButton = ImageTk.PhotoImage(resized)

    #create your back button
    Button(storeFrame,
    image=backButton,
    command= goBack,
    bg="#EDF7F7",
    relief="flat",
    borderwidth=0).pack(
        side="top",
        anchor="nw",
        pady=15,
        padx=30
    )

    #creating your header
    Label(storeFrame, 
    text="POWER-UP STORE",
    font=("Montserrat Extrabold", 36),
    fg="black",
    bg="#EDF7F7",
    justify="center"
    ).pack()

    #for the center frame
    centerFrame = Frame(storeFrame, bg="#EDF7F7")
    centerFrame.pack()

    saveMePic = Image.open("images/save me.png")
    eliminatePic = Image.open("images/eliminate.png")
    instantPic = Image.open("images/instant.png")
    hintPic = Image.open("images/hint.png")


    images = [saveMePic,eliminatePic,instantPic,hintPic] 
    imageList =[]

    wb = load_workbook("users database.xlsx")
    sheet = wb["Sheet1"]

    
    instantQuantity =25
    eliminateQuantity = 16
    hintQuantity = 12

    
    

    #setting image height
    for i in images:
        myHeight = 160
        aspectRatio = i.width/i.height
        newWidth = int(myHeight * aspectRatio)

        resized = i.resize((newWidth, myHeight))
        tk_image = ImageTk.PhotoImage(resized)
        imageList.append(tk_image)

    

    #for item name
    saveMe = Frame(centerFrame, bg= "white")
    #saveMe.grid(row = 0, column = 0, padx=15)
    saveMe.pack(side="left",  padx=15)

    #space below
    Label(saveMe, bg= "white").pack(pady=3)

    saveTitle = Label(saveMe,
    text= "SAVE ME",
    font= ("Montserrat Extrabold", 24),
    justify= "center",
    bg="white",
    fg="black")
    saveTitle.pack(pady=5)

    Label(saveMe,
    text="Maintain your progress and keep playing after getting a question wrong.",
    font= ("Montserrat", 10),
    justify= "center",
    bg="white",
    wraplength= 230,
    fg="black").pack(pady=2)

    Workbook = load_workbook("users database.xlsx")
    sheet = Workbook.active
    save = sheet[f"D{userIndex}"].value

    #Gift title
    smQuantity = Label(saveMe,
    text =f"In-stock: {save}",
    font= ("Montserrat", 16, "bold"),
    bg="white",
    fg="#1581B4"
    )
    smQuantity.pack()

    

    Label(saveMe, 
    bg ="white",
    image = imageList[0]).pack(pady=10)

    saveBuyButton = Button(saveMe,
    bg="#1581B4",
    fg="white",
    font=("Montserrat", 14, "bold"),
    justify="center",
    text="buy",
    pady=3,
    width=18
    
    )
    saveBuyButton.pack(pady=8,padx=20)

    #space below
    Label(saveMe, bg= "white").pack(pady=2)
    #****************************************************************************************************************************
    #Instant
    #for item name
    instant = Frame(centerFrame, bg= "white")
    instant.pack(side="left",  padx=15)
    #instant.grid(row = 0, column = 1, padx=15)

    #space below
    Label(instant, bg= "white").pack(pady=3)

    instantTitle = Label(instant,
    text= "INSTANT",
    font= ("Montserrat Extrabold", 24),
    justify= "center",
    bg="white",
    fg="black")
    instantTitle.pack(pady=5)

    Label(instant,
    text= "Provides the correct answer to the current question.",
    font= ("Montserrat", 10),
    justify= "center",
    bg="white",
    wraplength= 230,
    fg="black").pack(pady=2)

    #taking the amount value from the external databases
    Workbook = load_workbook("users database.xlsx")
    sheet = Workbook.active
    instantAmount = sheet[f"E{userIndex}"].value

    #Gift title
    instantQuantity = Label(instant,
    text =f"In-stock: {instantAmount}",
    font= ("Montserrat", 16, "bold"),
    bg="white",
    fg="#1581B4"
    )
    instantQuantity.pack()

    Label(instant, 
    bg ="white",
    image = imageList[2]).pack(pady=10)

    instantBuyButton = Button(instant,
    bg="#1581B4",
    fg="white",
    font=("Montserrat", 14, "bold"),
    justify="center",
    text="buy",
    pady=3,
    width=18
    
    )
    instantBuyButton.pack(pady=8,padx=20)

    #space below
    Label(instant, bg= "white").pack(pady=2)


    #**************************************************************************************************************************
    #Eliminate
    #for item name
    eliminate = Frame(centerFrame, bg= "white")
    eliminate.pack(side="left",  padx=15)
    #instant.grid(row = 0, column = 1, padx=15)

    #space below
    Label(eliminate, bg= "white").pack(pady=3)

    eliminateTitle = Label(eliminate,
    text= "ELIMINATE",
    font= ("Montserrat Extrabold", 24),
    justify= "center",
    bg="white",
    fg="black")
    eliminateTitle.pack(pady=5)

    Label(eliminate,
    text= "Eliminates two wrong answers from the current question.",
    font= ("Montserrat", 10),
    justify= "center",
    bg="white",
    wraplength= 230,
    fg="black").pack(pady=2)

    #taking the amount value from the external databases
    Workbook = load_workbook("users database.xlsx")
    sheet = Workbook.active
    eliminateAmount = sheet[f"F{userIndex}"].value

    #Gift title
    eliminateQuantity = Label(eliminate,
    text =f"In-stock: {eliminateAmount}",
    font= ("Montserrat", 16, "bold"),
    bg="white",
    fg="#1581B4"
    )
    eliminateQuantity.pack()

    Label(eliminate, 
    bg ="white",
    image = imageList[1]).pack(pady=10)

    eliminateBuyButton = Button(eliminate,
    bg="#1581B4",
    fg="white",
    font=("Montserrat", 14, "bold"),
    justify="center",
    text="buy",
    pady=3,
    width=18
    
    )
    eliminateBuyButton.pack(pady=8,padx=20)

    #space below
    Label(eliminate, bg= "white").pack(pady=2)
    #***************************************************************************************************************************
     #Eliminate
    #for item name
    hint = Frame(centerFrame, bg= "white")
    hint.pack(side="left",  padx=15)
    #instant.grid(row = 0, column = 1, padx=15)

    #space below
    Label(hint, bg= "white").pack(pady=3)

    hintTitle = Label(hint,
    text= "HINT",
    font= ("Montserrat Extrabold", 24),
    justify= "center",
    bg="white",
    fg="black")
    hintTitle.pack(pady=5)

    Label(hint,
    text= "Displays a hint on the current question.",
    font= ("Montserrat", 10),
    justify= "center",
    bg="white",
    wraplength= 230,
    fg="black").pack(pady=2)

    #taking the amount value from the external databases
    Workbook = load_workbook("users database.xlsx")
    sheet = Workbook.active
    hintAmount = sheet[f"G{userIndex}"].value

    #Gift title
    hintQuantity = Label(hint,
    text =f"In-stock: {hintAmount}",
    font= ("Montserrat", 16, "bold"),
    bg="white",
    fg="#1581B4"
    )
    hintQuantity.pack()

    Label(hint, 
    bg ="white",
    image = imageList[3]).pack(pady=10)

    hintBuyButton = Button(hint,
    bg="#1581B4",
    fg="white",
    font=("Montserrat", 14, "bold"),
    justify="center",
    text="buy",
    pady=3,
    width=18
    
    )
    hintBuyButton.pack(pady=8,padx=20)

    #space below
    Label(hint, bg= "white").pack(pady=2)

    buyButtons = [saveBuyButton, instantBuyButton, eliminateBuyButton, hintBuyButton]
    amountInStock = [smQuantity, instantQuantity, eliminateQuantity, hintQuantity]
    return storeFrame, amountInStock, buyButtons