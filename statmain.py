#import necessary libraries
from tkinter import *
from tkinter import ttk
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import load_workbook 
from PIL import Image, ImageTk
from tkinter import PhotoImage




def statgenerator(mockroot):
    def backbutton():
        print("success")
        
    lightblue="#ADD8E6"
    darkblue="#1581B4"
    root=Frame(mockroot,bg=lightblue)

    #load icon images
    stat_icon=Image.open("stat_icon.png")
    leaderboard_icon=Image.open("leaderboard_icon.png")
    backbutton_icon=Image.open("backbutton_icon.png")
    
    #resize the images based on locations
    stat_icon=stat_icon.resize((32,32))
    leaderboard_icon=leaderboard_icon.resize((32,32))
    backbutton_icon=backbutton_icon.resize((32,32))

    #return as a tkinter compatibel image
    stat_icon=ImageTk.PhotoImage(stat_icon)
    leaderboard_icon=ImageTk.PhotoImage(leaderboard_icon)
    backbutton_icon=ImageTk.PhotoImage(backbutton_icon)


    #windows initialization
    root.configure(bg=lightblue)
    root_frame=Frame(root,bg=lightblue)
    root_frame.pack(fill="x",expand=False)
    #root.title("Stat Page")
    title=Label(root_frame, text="QUIZ LARK", bg=lightblue,fg="black", font=("Comic Sans", 24, "bold"), highlightbackground="white")
    
    #define back button
    backbutton=Button(root_frame,image=backbutton_icon,bg=lightblue,command=backbutton)
    backbutton.image=backbutton_icon
    
    #display
    backbutton.pack(side=RIGHT)
    title.pack()

    #load user data base and fill in the necessary variables.
    file=load_workbook("users database.xlsx")
    sheet=file.active

    #left-top frame:
    leftframe=Frame(root,bg=lightblue)
    leftframe.pack(side=LEFT,fill="y")

    #statistics frame build up
    statframe =Frame(leftframe, bg=lightblue, relief="ridge")
    statframe.pack(fill="x")
    stat=Label(statframe, text="STATISTICS",image=stat_icon,compound=LEFT,bg=lightblue, fg=darkblue,  font=("Arial", 12, "bold"))
    stat.image=stat_icon
    stat.pack(anchor=NW)
    #user
    filedata=load_workbook("users database.xlsx")
    sheet=filedata.active

    # row variable to manipulate stat data output
    i=2

    #variables
    ttl_question=int(sheet[f"D{i}"].value)
    correct_ans=int(sheet[f"E{i}"].value)
    wrong_ans=int(sheet[f"F{i}"].value)
    game_played=int(sheet[f"G{i}"].value)
    games_won=int(sheet[f"H{i}"].value)
    streak=int(sheet[f"I{i}"].value)
    xp=int(sheet[f"J{i}"].value)

    #define the categories
    cat_a=int(sheet[f"K{i}"].value)
    cat_b=int(sheet[f"L{i}"].value)
    cat_c=int(sheet[f"M{i}"].value)
    cat_d=int(sheet[f"N{i}"].value)
    #selcetion to prevent logic error
    if ttl_question==0:
        correct_percent=0
        wrong_percent=0
    else:
        correct_percent=round((correct_ans*100)/ttl_question)
        wrong_percent=round((wrong_ans*100)/ttl_question)

    # Data List using nested lists...
    stats = [
        ["- Questions answered", ttl_question],
        ["- Correct answers", correct_ans],
        ["- Wrong answers", wrong_ans],
        ["- Correct percentage", correct_percent],
        ["- Wrong Percentage", wrong_percent],
        ["- Games played", game_played],
        ["- Games won", games_won],
        ["- Streak", streak],
        ["- XP(Experience points)", xp],
    ]
    #get userdata from database
    leaderboard_data = []

    #prevent None Error
    for i in range(2,12):
        if sheet[f"A{i}"].value !=None:
            leaderboard_data.append([sheet[f"A{i}"].value,sheet[f"J{i}"].value,sheet[f"I{i}"].value,sheet[f"H{i}"].value,sheet[f"C{i}"].value])
        else:
            break

    
    #sort the values using the xp key
    leaderboard_data.sort(key=lambda x: x[1],reverse=True)

    #join the datas again
    sorted_leader_data=leaderboard_data




    

    # Display the statistics in frame
    for text, value in stats:
        row_frame = Frame(statframe, bg="white")
        row_frame.pack(fill="x", padx=10, pady=3)

        label = Label(row_frame, text=text, font=("papyrus", 11,"bold"), bg="white", fg="black", anchor="w")
        label.pack(side="left")

        value_label = Label(row_frame, text=value, font=("Good Vibes", 11, "bold"), bg="white", fg="black", anchor="e")
        value_label.pack(side="right")


    #figure/pie size
    pieframe=Frame(root, bg="white", width=150,height=300000,relief="ridge")
    pieframe.pack(side="right",expand=True)
    figure=Figure(figsize=(5,6.2),dpi=80)
    subplot=figure.add_subplot(111)
    
    #selcetion to prevent logic error
    if cat_a ==0 and cat_b==0 and cat_c and cat_d:
        labels=["Fresh Player"]
        sizes=[100]
        colors=[lightblue]
        explode=[0]
    else:
        labels=['A','B','C','D']
        sizes=[cat_a,cat_b,cat_c,cat_d]
        colors=[lightblue,darkblue,"steelblue","powderblue"]
        explode=[0]

    subplot.pie(sizes,labels=labels,colors=colors,autopct='%1.1f%%')
    canvas=FigureCanvasTkAgg(figure,master=pieframe)
    canvas.draw()

    canvas.get_tk_widget().pack(fill="x")
    Label(pieframe,text="_________________________________________________________",bg="white",fg="black").pack(fill="x",expand=True)
    Label(pieframe,text=
    """
    Category A:    HISTORY
    Category B:    SCIENCE 
    Category C:    SPORTS
    Category D:    POP CULTURE
    """
    ,
    bg="white",fg="black",font=("Montserrat",12,"bold")).pack(anchor=NW)
    Label(pieframe,text="Top Categories",bg=darkblue,fg="white",font=("Arial", 14, "bold")).pack(fill="x",expand=True)
    
    # leaderboard frame
    leaderboard_frame =Frame(leftframe, bg=lightblue, padx=10, pady=10)
    leaderboard_frame.pack(side="left",fill="x")

    leaderdata=Label(leaderboard_frame, text="LEADERBOARDS", image=leaderboard_icon,compound=LEFT,font=("Arial", 12, "bold"), bg=lightblue)
    leaderdata.image=leaderboard_icon
    leaderdata.pack(anchor=NW)

    cols = ("Username", "XP", "Streak", "Games Won", "Name")
    view = ttk.Treeview(leaderboard_frame, columns=cols, show="headings")

    #fill in values using the for loop 
    for col in cols:
        view.heading(col, text=col)
        view.column(col, width=150)

    for row in sorted_leader_data:
        view.insert("", "end", values=row)

    view.pack(side=LEFT,expand=True)
    
    #return the root
    return root


